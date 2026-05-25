"""
Vietnam Scam Reviews – FREE NLP Analysis Pipeline
===================================================
Không cần API key, không tốn phí — dùng thư viện NLP miễn phí:

  • VADER  → sentiment score (được thiết kế cho social media / review text)
  • Keyword matching → phân loại scam category + trích xuất thành phố
  • Severity scoring → tính dựa trên từ khoá mức độ + tài chính
  • Risk Score formula: Severity×40% + Sentiment×30% + Frequency×30%

Cài đặt (một lần duy nhất):
  pip install vaderSentiment pandas openpyxl

Chạy:
  python analyze_free.py
"""

import re
import pandas as pd
from collections import Counter, defaultdict
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── CONFIG ─────────────────────────────────────────────────────────────────────

INPUT_CSV    = "vietnam_scam_reviews.csv"
OUTPUT_XLSX  = "vietnam_scam_analysis.xlsx"
OUTPUT_CSV   = "vietnam_scam_analysis.csv"

# ── CLASSIFICATION RULES ───────────────────────────────────────────────────────
# Mỗi category có list từ khoá, xếp theo độ ưu tiên (category đầu tiên match = win)

CATEGORY_RULES = [
    ("safety", [
        "assault", "attacked", "drugged", "spiked", "blackout", "unconscious",
        "robbery", "robbed", "mugged", "threatened", "violence", "rape",
        "drink spiking", "dangerous", "unsafe", "molested", "harassed sexually"
    ]),
    ("digital_app", [
        "qr code", "qr scam", "fake app", "fake website", "phishing",
        "transfer scam", "wrong account", "fake payment", "digital wallet",
        "online scam", "internet scam", "fake grab app", "wrong number"
    ]),
    ("money_exchange", [
        "money exchange", "currency exchange", "exchange rate", "fake bill",
        "counterfeit", "fake note", "wrong change", "short change",
        "exchange booth", "exchange shop", "money changer", "fake money",
        "vnd", "dong", "cash scam", "atm scam", "atm fraud"
    ]),
    ("taxi_grab", [
        "taxi", "grab", "driver", "motorbike", "xe om", "xe ôm",
        "tuktuk", "tuk tuk", "cyclo", "uber", "ride", "fare",
        "meter", "overcharged taxi", "fake grab", "grab scam",
        "no meter", "rigged meter", "wrong route", "airport taxi"
    ]),
    ("tour_attraction", [
        "tour", "boat trip", "halong", "ha long", "cruise",
        "guide", "tourist guide", "fake ticket", "entrance fee",
        "attraction", "temple", "museum", "day trip", "travel agency",
        "booking tour", "tour scam", "overpriced tour", "visa scam"
    ]),
    ("restaurant_bar", [
        "restaurant", "bar", "cafe", "food", "drink", "beer",
        "menu", "bill", "overcharged food", "fake menu", "hidden charge",
        "street food", "pho", "banh mi", "mojito", "cocktail",
        "price not listed", "no price", "charged extra", "inflated price"
    ]),
    ("accommodation", [
        "hotel", "hostel", "guesthouse", "airbnb", "booking", "room",
        "accommodation", "check in", "check out", "deposit", "refund",
        "fake hotel", "wrong hotel", "extra charge hotel", "hidden fee hotel"
    ]),
    ("shopping", [
        "shop", "market", "souvenir", "buy", "purchase", "silk",
        "lacquer", "jewelry", "gem", "stone", "jade", "fake goods",
        "overpriced", "return policy", "no refund", "counterfeit goods",
        "haggle", "vendor", "ben thanh", "night market"
    ]),
    ("street_vendor", [
        "street vendor", "shoe shine", "shoe repair", "flower seller",
        "fruit vendor", "hat scam", "put hat on", "unsolicited",
        "uninvited", "forced service", "weight fruit", "photo scam"
    ]),
    ("other", [])  # fallback
]

# Từ khoá xác định thành phố
CITY_RULES = [
    ("hcmc",      ["hcmc", "ho chi minh", "saigon", "ho-chi-minh", "district 1",
                   "ben thanh", "bui vien", "pham ngu lao", "tan son nhat", "sgn"]),
    ("hanoi",     ["hanoi", "ha noi", "ha-noi", "old quarter", "hoan kiem",
                   "noi bai", "west lake", "ho tay"]),
    ("hoi_an",    ["hoi an", "hoian", "hoi-an", "ancient town hoi"]),
    ("da_nang",   ["da nang", "danang", "da-nang", "my khe", "han river", "marble mountain"]),
    ("nha_trang", ["nha trang", "nhatrang", "nha-trang"]),
    ("phu_quoc",  ["phu quoc", "phuquoc", "phu-quoc"]),
    ("sapa",      ["sapa", "sa pa", "sa-pa", "fansipan"]),
    ("ha_long",   ["halong", "ha long", "ha-long", "halong bay", "ha long bay"]),
]

# Từ khoá severity (mức độ nghiêm trọng)
SEVERITY_HIGH = [
    "assault", "attacked", "robbed", "dangerous", "unconscious", "drugged",
    "spiked", "blackout", "lost everything", "stolen", "pickpocket",
    "large amount", "hundreds of dollars", "lot of money", "significant",
    "traumatic", "terrifying", "worst experience"
]
SEVERITY_MED_HIGH = [
    "overcharged", "ripped off", "scammed", "cheated", "fraud", "fake",
    "angry", "furious", "terrible", "horrible", "awful", "never again",
    "warned everyone", "lost money", "wasted", "regret"
]
SEVERITY_MED = [
    "annoying", "frustrating", "disappointed", "tricked", "misled",
    "confused", "uncomfortable", "unpleasant", "beware", "warning"
]
SEVERITY_LOW = [
    "minor", "small", "little", "slight", "bit", "not much",
    "manageable", "ok in the end", "resolved", "got money back"
]

# Từ khoá mất tiền tài chính
LOSS_HIGH    = ["hundreds", "million", "large sum", "lot of money", "significant amount",
                "all my money", "entire budget", "2 million", "3 million", "5 million"]
LOSS_MED     = ["thousand", "500k", "overcharged", "extra charge", "double", "triple"]
LOSS_LOW     = ["small amount", "few dollars", "minor", "little money", "50k", "100k"]
LOSS_NONE    = ["no money lost", "got refund", "money back", "resolved", "avoided"]

PREVENTION_TIPS = {
    "taxi_grab":      "Always verify price in Grab app before boarding",
    "money_exchange": "Exchange only at licensed banks or official ATMs",
    "restaurant_bar": "Ask for English menu with prices before ordering",
    "tour_attraction":"Book tours only through certified travel agencies",
    "shopping":       "Compare prices at multiple shops, always bargain",
    "accommodation":  "Book only via verified platforms like Booking.com",
    "street_vendor":  "Politely decline any unsolicited service offers",
    "digital_app":    "Verify recipient account before confirming transfer",
    "safety":         "Never leave your drink unattended in bars/clubs",
    "other":          "Research common local scams before visiting",
}

SCAM_METHODS = {
    "taxi_grab":      "Meter tampering or fake Grab badge overcharge",
    "money_exchange": "Wrong change or counterfeit bills given",
    "restaurant_bar": "Hidden charges or bait menu pricing",
    "tour_attraction":"Unlicensed operators with fake tickets",
    "shopping":       "Counterfeit goods sold at premium prices",
    "accommodation":  "Bait-and-switch room or hidden fees",
    "street_vendor":  "Uninvited service then forced payment",
    "digital_app":    "QR code redirects to fraudulent payment",
    "safety":         "Drink spiking targeting solo tourists",
    "other":          "Various deceptive practices on tourists",
}

# ── NLP ENGINE ─────────────────────────────────────────────────────────────────

vader = SentimentIntensityAnalyzer()

# Thêm từ khoá scam vào VADER lexicon để tăng độ chính xác
SCAM_BOOSTERS = {
    "scammed": -3.0, "ripped off": -3.0, "overcharged": -2.5,
    "cheated": -2.8, "fraud": -3.0, "fake": -2.0, "beware": -1.5,
    "warning": -1.2, "avoid": -1.5, "terrible": -2.5, "awful": -2.5,
    "trap": -2.0, "dishonest": -2.5, "con": -2.0, "swindled": -3.0,
    "pickpocketed": -3.0, "stolen": -3.0, "robbery": -3.5,
}
vader.lexicon.update(SCAM_BOOSTERS)


def classify_category(text: str) -> str:
    text_lower = text.lower()
    for category, keywords in CATEGORY_RULES:
        if any(kw in text_lower for kw in keywords):
            return category
    return "other"


def extract_city(text: str) -> str:
    text_lower = text.lower()
    for city, keywords in CITY_RULES:
        if any(kw in text_lower for kw in keywords):
            return city
    return "unknown"


def is_vietnam_related(text: str) -> bool:
    vn_kw = [
        "vietnam", "viet nam", "vietnamese", "viet", "hcmc", "ho chi minh",
        "saigon", "hanoi", "ha noi", "hoi an", "da nang", "nha trang",
        "phu quoc", "sapa", "halong", "grab", "vnd", "dong currency"
    ]
    return any(kw in text.lower() for kw in vn_kw)


def compute_sentiment(text: str) -> float:
    """
    VADER trả compound score [-1, +1].
    Chuyển sang sentiment_score [0, 10]:
      -1.0 compound → 10 (cực kỳ tiêu cực)
      +1.0 compound → 0  (rất tích cực)
    """
    scores = vader.polarity_scores(text)
    compound = scores["compound"]
    # Linear transform: compound -1→10, +1→0
    sentiment = round((1 - compound) / 2 * 10, 1)
    return max(0.0, min(10.0, sentiment))


def compute_severity(text: str) -> int:
    """
    Tính severity 1–5 dựa trên từ khoá mức độ nghiêm trọng.
    """
    text_lower = text.lower()
    if any(kw in text_lower for kw in SEVERITY_HIGH):
        base = 5
    elif any(kw in text_lower for kw in SEVERITY_MED_HIGH):
        base = 4
    elif any(kw in text_lower for kw in SEVERITY_MED):
        base = 3
    elif any(kw in text_lower for kw in SEVERITY_LOW):
        base = 2
    else:
        base = 3  # default: trung bình nếu là scam post

    # Điều chỉnh dựa trên số tiền nhắc đến
    if any(kw in text_lower for kw in ["$100", "$200", "$500", "2 million", "5 million"]):
        base = min(5, base + 1)
    if any(kw in text_lower for kw in ["$5", "$10", "small", "minor"]):
        base = max(1, base - 1)

    return base


def compute_financial_loss(text: str) -> str:
    text_lower = text.lower()
    if any(kw in text_lower for kw in LOSS_NONE):
        return "none"
    if any(kw in text_lower for kw in LOSS_HIGH):
        return "high"
    if any(kw in text_lower for kw in LOSS_MED):
        return "medium"
    if any(kw in text_lower for kw in LOSS_LOW):
        return "low"
    return "unknown"


def compute_confidence(text: str, category: str, city: str) -> float:
    """
    Ước tính confidence dựa trên:
    - Độ dài bài (nhiều từ = nhiều tín hiệu = confidence cao hơn)
    - Category có match keyword không (vs fallback "other")
    - City có identify được không
    """
    score = 0.5
    word_count = len(text.split())
    if word_count > 100: score += 0.2
    elif word_count > 50: score += 0.1
    if category != "other": score += 0.15
    if city != "unknown":   score += 0.15
    return round(min(0.99, score), 2)


def analyze_review(row: pd.Series) -> dict:
    """Phân tích một review, trả về dict các trường AI."""
    title = str(row.get("title", ""))
    text  = str(row.get("review_text", ""))
    full  = f"{title}. {text}"

    is_vn    = is_vietnam_related(full)
    category = classify_category(full)
    city     = extract_city(full)
    sentiment= compute_sentiment(full)
    severity = compute_severity(full)
    fin_loss = compute_financial_loss(full)
    conf     = compute_confidence(full, category, city)

    return {
        "is_vietnam":      is_vn,
        "scam_category":   category,
        "city":            city,
        "sentiment_score": sentiment,
        "severity":        severity,
        "financial_loss":  fin_loss,
        "scam_method":     SCAM_METHODS.get(category, "Unknown method"),
        "tourist_tip":     PREVENTION_TIPS.get(category, "Be cautious"),
        "confidence":      conf,
    }


# ── RISK SCORE ─────────────────────────────────────────────────────────────────

def calculate_risk_scores(df: pd.DataFrame):
    vn = df[df["is_vietnam"] == True].copy()

    def risk_formula(items_df):
        sev  = items_df["severity"].mean()
        sent = items_df["sentiment_score"].mean()
        n    = len(items_df)
        freq = min(n / 3, 10)
        return round((sev/5*40) + (sent/10*30) + (freq/10*30), 1)

    # Risk by City
    city_rows = []
    for city, grp in vn[vn["city"] != "unknown"].groupby("city"):
        top_cat = grp["scam_category"].mode()[0] if len(grp) > 0 else "other"
        city_rows.append({
            "city":             city,
            "total_reviews":    len(grp),
            "avg_severity":     round(grp["severity"].mean(), 2),
            "avg_sentiment":    round(grp["sentiment_score"].mean(), 2),
            "high_loss_pct":    round(grp["financial_loss"].isin(["high","medium"]).mean()*100, 1),
            "risk_score":       risk_formula(grp),
            "top_scam_type":    top_cat,
            "prevention_tip":   PREVENTION_TIPS.get(top_cat, "Be cautious"),
        })
    city_df = pd.DataFrame(city_rows).sort_values("risk_score", ascending=False) \
              if city_rows else pd.DataFrame()

    # Risk by Category
    cat_rows = []
    for cat, grp in vn.groupby("scam_category"):
        top_city = grp["city"].mode()[0] if len(grp) > 0 else "unknown"
        cat_rows.append({
            "scam_category":    cat,
            "total_reviews":    len(grp),
            "avg_severity":     round(grp["severity"].mean(), 2),
            "avg_sentiment":    round(grp["sentiment_score"].mean(), 2),
            "high_loss_pct":    round(grp["financial_loss"].isin(["high","medium"]).mean()*100, 1),
            "risk_score":       risk_formula(grp),
            "top_city":         top_city,
            "typical_method":   SCAM_METHODS.get(cat, "Unknown"),
            "prevention_tip":   PREVENTION_TIPS.get(cat, "Be cautious"),
        })
    cat_df = pd.DataFrame(cat_rows).sort_values("risk_score", ascending=False) \
             if cat_rows else pd.DataFrame()

    return city_df, cat_df


# ── EXCEL STYLING ──────────────────────────────────────────────────────────────

DARK_BLUE  = "1F4E79"; MED_BLUE  = "2E75B6"; LIGHT_BLUE  = "D6E4F0"
LIGHT_RED  = "FADBD8"; LIGHT_AMBER = "FFF2CC"; LIGHT_GREEN = "D9EAD3"
WHITE = "FFFFFF"

def tb():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg=DARK_BLUE):
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = tb()

def dat(cell, alt=False, wrap=False, center=False):
    cell.font      = Font(name="Arial", size=9)
    cell.fill      = PatternFill("solid", fgColor=LIGHT_BLUE if alt else WHITE)
    cell.alignment = Alignment(
        vertical="top", wrap_text=wrap,
        horizontal="center" if center else "left"
    )
    cell.border = tb()

def risk_fill(v):
    try:
        v = float(v)
        if v >= 70: return LIGHT_RED
        if v >= 45: return LIGHT_AMBER
        return LIGHT_GREEN
    except: return WHITE

def sev_fill(v):
    try:
        v = float(v)
        if v >= 4: return LIGHT_RED
        if v >= 3: return LIGHT_AMBER
        return LIGHT_GREEN
    except: return WHITE


def export_excel(df: pd.DataFrame, city_df: pd.DataFrame,
                 cat_df: pd.DataFrame, path: str):

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer,      sheet_name="📋 All Reviews",      index=False)
        if not city_df.empty:
            city_df.to_excel(writer, sheet_name="🏙 Risk by City",     index=False)
        if not cat_df.empty:
            cat_df.to_excel(writer,  sheet_name="⚠ Risk by Category", index=False)

    wb = load_workbook(path)

    # ── Sheet 1: All Reviews ───────────────────────────────────────────────────
    ws = wb["📋 All Reviews"]
    col_w = {"A":10,"B":14,"C":12,"D":40,"E":55,"F":8,"G":10,
             "H":16,"I":14,"J":14,"K":14,"L":8,"M":10,"N":35,"O":35,"P":10}
    for c, w in col_w.items():
        if c in ws.column_dimensions:
            ws.column_dimensions[c].width = w
    ws.row_dimensions[1].height = 28
    for ci in range(1, ws.max_column+1):
        hdr(ws.cell(1, ci))

    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    sev_ci  = (hdrs.index("severity")+1)  if "severity"  in hdrs else None
    sent_ci = (hdrs.index("sentiment_score")+1) if "sentiment_score" in hdrs else None
    wrap_ci = {hdrs.index(h)+1 for h in ["title","review_text","scam_method","tourist_tip"]
               if h in hdrs}

    for ri in range(2, ws.max_row+1):
        ws.row_dimensions[ri].height = 48
        alt = ri % 2 == 0
        for ci in range(1, ws.max_column+1):
            dat(ws.cell(ri, ci), alt=alt, wrap=(ci in wrap_ci))
        if sev_ci:
            c = ws.cell(ri, sev_ci)
            c.fill = PatternFill("solid", fgColor=sev_fill(c.value))
            c.font = Font(name="Arial", size=10, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheets 2 & 3: Risk tables ──────────────────────────────────────────────
    for sname in [s.title for s in wb.worksheets if "Risk" in s.title]:
        ws2 = wb[sname]
        for c in [get_column_letter(i) for i in range(1, ws2.max_column+1)]:
            ws2.column_dimensions[c].width = 20
        ws2.row_dimensions[1].height = 28
        for ci in range(1, ws2.max_column+1):
            hdr(ws2.cell(1, ci))
        for ri in range(2, ws2.max_row+1):
            ws2.row_dimensions[ri].height = 26
            for ci in range(1, ws2.max_column+1):
                c = ws2.cell(ri, ci)
                c.font      = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = tb()
            # risk_score = col 6
            rc = ws2.cell(ri, 6)
            try:
                rc.fill = PatternFill("solid", fgColor=risk_fill(rc.value))
                rc.font = Font(name="Arial", size=10, bold=True)
            except: pass
        try:
            ws2.conditional_formatting.add(
                f"F2:F{ws2.max_row}",
                ColorScaleRule(start_type="min",  start_color="63BE7B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max",    end_color="F8696B"))
        except: pass

    # ── Sheet 4: Dashboard ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("📊 Dashboard")
    ws4.sheet_view.showGridLines = False
    for c, w in [("A",3),("B",32),("C",20),("D",32),("E",20)]:
        ws4.column_dimensions[c].width = w

    def mkh(row, text, bg=DARK_BLUE):
        c = ws4.cell(row, 2, f"  {text}")
        c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        ws4.merge_cells(f"B{row}:E{row}")
        ws4.row_dimensions[row].height = 28

    def mkrow(row, label, value, lbg=MED_BLUE, vbg=LIGHT_BLUE):
        lc = ws4.cell(row, 2, label)
        lc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        lc.fill = PatternFill("solid", fgColor=lbg)
        lc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        lc.border = tb()
        vc = ws4.cell(row, 3, value)
        vc.font = Font(name="Arial", size=10)
        vc.fill = PatternFill("solid", fgColor=vbg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = tb()
        ws4.row_dimensions[row].height = 24

    # Title
    t = ws4.cell(2, 2, "📊  Vietnam Travel Risk Dashboard")
    t.font = Font(name="Arial", bold=True, size=16, color=DARK_BLUE)
    ws4.merge_cells("B2:E2"); ws4.row_dimensions[2].height = 40

    sub = ws4.cell(3, 2, "Free NLP Analysis · VADER Sentiment + Keyword Classification · 321 Reviews")
    sub.font = Font(name="Arial", size=10, color="666666")
    ws4.merge_cells("B3:E3"); ws4.row_dimensions[3].height = 20

    vn_count = int(df["is_vietnam"].sum())
    mkh(5, "DATA OVERVIEW")
    mkrow(6,  "Tổng số reviews",                    len(df))
    mkrow(7,  "Vietnam-related reviews",             vn_count)
    mkrow(8,  "Avg severity score (1–5)",            round(df["severity"].mean(), 2))
    mkrow(9,  "Avg negative sentiment (0–10)",       round(df["sentiment_score"].mean(), 2))
    mkrow(10, "High financial loss cases",           int((df["financial_loss"]=="high").sum()))
    mkrow(11, "Phương pháp phân tích",               "VADER Sentiment + Keyword NLP (free)")

    mkh(13, "TOP RISK CITIES  (Risk Score 0–100)")
    for i, (_, r) in enumerate(city_df.head(5).iterrows()):
        vbg = LIGHT_RED if r["risk_score"] >= 70 else LIGHT_AMBER
        mkrow(14+i,
              f"#{i+1}  {r['city'].replace('_',' ').title()}",
              f"Risk: {r['risk_score']}  |  {r['total_reviews']} reviews",
              lbg="2E75B6" if i==0 else "4A90D9", vbg=vbg)

    mkh(20, "TOP SCAM CATEGORIES  (Risk Score 0–100)")
    for i, (_, r) in enumerate(cat_df.head(5).iterrows()):
        vbg = LIGHT_RED if r["risk_score"] >= 70 else LIGHT_AMBER
        mkrow(21+i,
              f"#{i+1}  {r['scam_category'].replace('_',' ').title()}",
              f"Risk: {r['risk_score']}  |  {r['total_reviews']} reviews",
              lbg="2E75B6" if i==0 else "4A90D9", vbg=vbg)

    mkh(27, "RISK SCORE FORMULA", bg="1E5631")
    note = ws4.cell(28, 2,
        "Risk Score = Severity × 40%  +  Sentiment × 30%  +  Frequency × 30%")
    note.font  = Font(name="Arial", size=10, color=DARK_BLUE, bold=True, italic=True)
    note.fill  = PatternFill("solid", fgColor="EBF3FB")
    ws4.merge_cells("B28:E28")
    note.border = Border(left=Side(style="medium", color=MED_BLUE),
                         right=Side(style="thin"), top=Side(style="thin"),
                         bottom=Side(style="thin"))
    ws4.row_dimensions[28].height = 28

    note2 = ws4.cell(29, 2,
        "Severity: keyword-based 1–5  ·  Sentiment: VADER compound score  ·  Frequency: log-normalized count")
    note2.font  = Font(name="Arial", size=9, color="555555", italic=True)
    note2.fill  = PatternFill("solid", fgColor="F5F5F5")
    ws4.merge_cells("B29:E29")
    ws4.row_dimensions[29].height = 22

    wb.save(path)
    print(f"  ✓ Excel saved: {path}")


# ── MAIN ────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  Vietnam Scam Reviews – FREE NLP Analysis (no API needed)")
    print("=" * 65)

    df = pd.read_csv(INPUT_CSV)
    print(f"\n📂 Loaded {len(df)} reviews")

    print("\n🔍 Phân tích NLP (VADER + keyword matching)...")
    results = []
    for i, (_, row) in enumerate(df.iterrows()):
        results.append(analyze_review(row))
        if (i+1) % 50 == 0:
            print(f"   {i+1}/{len(df)} done...")

    result_df = pd.DataFrame(results)
    for col in result_df.columns:
        df[col] = result_df[col].values

    print(f"\n✅ Phân tích xong {len(df)} reviews")
    print(f"   Vietnam-related: {df['is_vietnam'].sum()}")
    print(f"   Avg sentiment:   {df['sentiment_score'].mean():.2f} / 10")
    print(f"   Avg severity:    {df['severity'].mean():.2f} / 5")

    print("\n📊 Tính Risk Score...")
    city_df, cat_df = calculate_risk_scores(df)

    print("\n🏙  Risk by City:")
    if not city_df.empty:
        print(city_df[["city","total_reviews","risk_score","top_scam_type"]]
              .to_string(index=False))

    print("\n⚠  Risk by Category:")
    if not cat_df.empty:
        print(cat_df[["scam_category","total_reviews","risk_score","top_city"]]
              .to_string(index=False))

    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"\n  ✓ CSV saved: {OUTPUT_CSV}")

    print("\n📝 Xuất Excel (4 sheets)...")
    export_excel(df, city_df, cat_df, OUTPUT_XLSX)

    print("\n✅ Hoàn thành — không tốn 1 đồng API!")
    print(f"   → {OUTPUT_XLSX}")
    print(f"   → {OUTPUT_CSV}")
    print("=" * 65)

if __name__ == "__main__":
    main()